"""
MEST-C component prediction script
Usage: python -m 3classification.MEST_C.predict --task M --config config.yaml --input_excel path.xlsx --output_csv path.csv
"""
from __future__ import annotations

import argparse
import os
import re

import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
from torch.utils.data import Dataset, DataLoader
from tqdm import tqdm

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import (
    roc_curve, auc, confusion_matrix,
    classification_report, accuracy_score, f1_score, roc_auc_score,
)
from sklearn.preprocessing import label_binarize

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from utils import plot_multiclass_roc, plot_confusion_matrix

from ..config import load_config, get_device
from ..models import GaanClassifier
from .data import TASK_CONFIG


# ============================================================
#  Prediction Dataset
# ============================================================
class MESTCPredictionDataset(Dataset):
    def __init__(self, excel_path: str, pt_dir: str, id_column: str = "ID"):
        df = pd.read_excel(excel_path)
        if id_column not in df.columns:
            raise ValueError(f"Column '{id_column}' not found. Available: {list(df.columns)}")
        self.df = df
        self.id_column = id_column
        self.pt_dir = pt_dir
        self.ids = df[id_column].astype(str).tolist()
        print(f"[OK] Predictable samples: {len(self.ids)}")

    def __len__(self):
        return len(self.ids)

    def __getitem__(self, idx):
        sample_id = self.ids[idx]
        pt_path = os.path.join(self.pt_dir, f"{sample_id}.pt")
        bag = torch.load(pt_path, weights_only=True)
        if isinstance(bag, dict) and "tokens" in bag:
            bag = bag["tokens"]
        return bag, sample_id


def collate_fn(batch):
    bags = [item[0] for item in batch]
    ids = [item[1] for item in batch]
    return bags, ids


# ============================================================
#  Model Loading
# ============================================================
def clean_state_dict(state_dict):
    new_state_dict = {}
    for k, v in state_dict.items():
        if k.startswith("module."):
            k = k.replace("module.", "", 1)
        new_state_dict[k] = v
    return new_state_dict


def load_gaan_model(model_path, device, num_classes=6):
    model = GaanClassifier(num_classes=num_classes).to(device)
    ckpt = torch.load(model_path, map_location=device)
    if isinstance(ckpt, dict):
        state_dict = ckpt.get("state_dict", ckpt.get("model_state_dict", ckpt))
    else:
        raise TypeError(f"Unknown checkpoint format: {type(ckpt)}")
    model.load_state_dict(clean_state_dict(state_dict))
    model.eval()
    print(f"[OK] Loaded GAAN model: {model_path}")
    return model


# ============================================================
#  Predict
# ============================================================
def predict(
    input_excel: str,
    pt_dir: str,
    model_path: str,
    output_csv: str,
    task: str,
    id_column: str = "ID",
    device: str = "0",
):
    _, _, num_classes = TASK_CONFIG[task]
    class_names = [f"{task}{i}" for i in range(num_classes)]

    if device.isdigit():
        dev = torch.device(f"cuda:{device}")
    else:
        dev = torch.device(device)

    # Load model
    model = GaanClassifier(num_classes=num_classes).to(dev)
    ckpt = torch.load(model_path, map_location=dev, weights_only=False)
    model.load_state_dict(clean_state_dict(ckpt))
    model.eval()
    print(f"[OK] Loaded {task} model: {model_path} (classes={num_classes})")

    # Dataset
    dataset = MESTCPredictionDataset(input_excel, pt_dir, id_column)
    loader = DataLoader(dataset, batch_size=1, shuffle=False, collate_fn=collate_fn)

    # Predict
    results = []
    with torch.no_grad():
        for bags, sample_ids in tqdm(loader, desc=f"Predicting {task}"):
            bags = [b.to(dev) for b in bags]
            logits = model(bags)
            probs = F.softmax(logits, dim=1).cpu().numpy()[0]
            pred = int(np.argmax(probs))

            row = {id_column: sample_ids[0], f"{task}_pred": pred}
            for c in range(num_classes):
                row[f"{task}{c}_prob"] = float(probs[c])
            results.append(row)

    result_df = pd.DataFrame(results)

    # Merge with original Excel
    original_df = pd.read_excel(input_excel)
    result_df = pd.merge(original_df, result_df, on=id_column)

    # Save CSV
    os.makedirs(os.path.dirname(output_csv) or ".", exist_ok=True)
    result_df.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"[OK] {task} predictions saved to: {output_csv}")

    # Save Excel (probability document)
    excel_path = output_csv.replace(".csv", "_result.xlsx")
    result_df.to_excel(excel_path, index=False)
    print(f"[OK] Result document saved to: {excel_path}")

    # Auto-evaluate if true labels exist
    label_col = f"{task}" if task in result_df.columns else None
    if label_col and result_df[label_col].notna().any():
        true_vals = pd.to_numeric(result_df[label_col], errors="coerce")
        valid = true_vals.notna()
        if valid.any():
            y_true = true_vals[valid].astype(int).values
            y_pred = result_df.loc[valid, f"{task}_pred"].astype(int).values
            y_score = result_df.loc[valid, [f"{task}{c}_prob" for c in range(num_classes)]].values

            print(f"\n[{task}] Classification Report:")
            print(classification_report(y_true, y_pred, labels=list(range(num_classes)),
                                        target_names=class_names, digits=4, zero_division=0))

            # Auto-generate plots
            save_dir = os.path.dirname(output_csv) or "."

            # 1. ROC Curve
            if num_classes == 2:
                from sklearn.metrics import roc_curve as roc_f, auc as roc_a
                pos_probs = y_score[:, 1]
                fpr, tpr, _ = roc_f(y_true, pos_probs)
                roc_auc_val = roc_a(fpr, tpr)
                plt.figure(figsize=(6, 5))
                plt.plot(fpr, tpr, lw=2, label=f"AUC = {roc_auc_val:.3f}")
                plt.plot([0, 1], [0, 1], "--", color="gray")
                plt.xlabel("False Positive Rate")
                plt.ylabel("True Positive Rate")
                plt.title(f"{task} Component - ROC Curve")
                plt.legend(loc="lower right")
                plt.tight_layout()
                plt.savefig(os.path.join(save_dir, f"{task}_roc.png"), dpi=300)
                plt.close()
            else:
                plot_multiclass_roc(y_true, y_score, class_names,
                                    os.path.join(save_dir, f"{task}_roc.png"))

            # 2. Confusion Matrix
            plot_confusion_matrix(y_true, y_pred, class_names,
                                  os.path.join(save_dir, f"{task}_confusion_matrix.png"))

            # 3. Summary xlsx with bootstrap CI
            _save_summary_xlsx(save_dir, class_names, y_true, y_pred, y_score, task)

            print(f"[OK] Evaluation plots + summary saved to: {save_dir}")

    return result_df


# ============================================================
#  Summary xlsx with bootstrap CI
# ============================================================
def _save_summary_xlsx(output_dir, class_names, y_true, y_pred, y_score, task_name, n_boot=1000):
    """rows=Metrics, columns=Classes, format: value [95% CI]"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    SEED = 42
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_score = np.array(y_score)
    n_classes = len(class_names)

    def point_ci_text(point, boot_vals):
        try:
            point = float(point)
        except Exception:
            return ""
        if np.isnan(point):
            return ""
        clean = [v for v in boot_vals if not np.isnan(v)]
        if len(clean) == 0:
            return f"{point:.4f}"
        low = np.percentile(clean, 2.5)
        high = np.percentile(clean, 97.5)
        return f"{point:.4f} [{low:.4f}-{high:.4f}]"

    def calc_metrics(yt, yp, yp_prob):
        cm_bin = confusion_matrix(yt, yp, labels=[0, 1])
        tn, fp, fn, tp = cm_bin.ravel()
        sens = tp / (tp + fn) if (tp + fn) > 0 else np.nan
        spec = tn / (tn + fp) if (tn + fp) > 0 else np.nan
        ppv = tp / (tp + fp) if (tp + fp) > 0 else np.nan
        npv = tn / (tn + fn) if (tn + fn) > 0 else np.nan
        f1 = f1_score(yt, yp, zero_division=0)
        try:
            a = roc_auc_score(yt, yp_prob) if len(np.unique(yt)) >= 2 else np.nan
        except Exception:
            a = np.nan
        return {'Sensitivity': sens, 'PPV (Precision)': ppv, 'AUC': a, 'F1-Score': f1, 'Specificity': spec, 'NPV': npv}

    raw_class = {}
    for i in range(n_classes):
        raw_class[i] = calc_metrics((y_true == i).astype(int), (y_pred == i).astype(int), y_score[:, i])

    rng = np.random.RandomState(SEED)
    n = len(y_true)
    boot_class = {i: {k: [] for k in ['Sensitivity', 'PPV (Precision)', 'AUC', 'F1-Score', 'Specificity', 'NPV']} for i in range(n_classes)}
    for _ in range(n_boot):
        idx = rng.randint(0, n, n)
        bt, bp, bs = y_true[idx], y_pred[idx], y_score[idx]
        for i in range(n_classes):
            m = calc_metrics((bt == i).astype(int), (bp == i).astype(int), bs[:, i])
            for k in boot_class[i]:
                boot_class[i][k].append(m[k])

    wb = Workbook()
    ws = wb.active
    ws.title = "具体数据"

    dark_blue, mid_blue, light_grid = "2F5597", "8EA9DB", "D9E2F3"
    title_fill = PatternFill("solid", fgColor=dark_blue)
    header_fill = PatternFill("solid", fgColor=mid_blue)
    left_fill = PatternFill("solid", fgColor=mid_blue)
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=light_grid)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    metric_rows = ["Sensitivity", "PPV", "AUC", "F1-score", "Specificity", "NPV"]
    metric_keys = ["Sensitivity", "PPV (Precision)", "AUC", "F1-Score", "Specificity", "NPV"]
    headers = ["Metrics"] + class_names

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).value = f"{task_name} Component Results"
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = white_font
        c.alignment = center
        c.border = border

    for row_i, (display_name, metric_key) in enumerate(zip(metric_rows, metric_keys)):
        row = 3 + row_i
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=1).fill = left_fill
        ws.cell(row=row, column=1).font = white_font
        ws.cell(row=row, column=1).alignment = center
        for col_i, cls_name in enumerate(class_names):
            point = raw_class[col_i].get(metric_key, np.nan)
            vals = boot_class[col_i].get(metric_key, [])
            ws.cell(row=row, column=2 + col_i, value=point_ci_text(point, vals)).font = body_font
            ws.cell(row=row, column=2 + col_i).alignment = center
            ws.cell(row=row, column=2 + col_i).border = border

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22

    xlsx_path = os.path.join(output_dir, f"{task_name}_具体数据.xlsx")
    wb.save(xlsx_path)
    print(f"[OK] Summary xlsx saved to: {xlsx_path}")


# ============================================================
#  CLI
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="MEST-C component prediction")
    parser.add_argument("--task", type=str, required=True, choices=["M", "E", "S", "T", "C"])
    parser.add_argument("--config", type=str, default=None, help="Path to config.yaml")
    parser.add_argument("--input_excel", type=str, required=True)
    parser.add_argument("--output_csv", type=str, required=True)
    parser.add_argument("--pt_dir", type=str, required=True, help="Directory of .pt feature files")
    parser.add_argument("--model_path", type=str, required=True, help="Model checkpoint path")
    parser.add_argument("--id_column", type=str, default="ID")
    parser.add_argument("--device", type=str, default="0")
    args = parser.parse_args()

    predict(
        input_excel=args.input_excel, pt_dir=args.pt_dir,
        model_path=args.model_path, output_csv=args.output_csv,
        task=args.task, id_column=args.id_column, device=args.device,
    )
