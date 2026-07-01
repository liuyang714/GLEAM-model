"""
Shared utilities: random seed, loss functions, ROC plotting, evaluation metrics
Includes macaron academic style plotting from original evaluation script.
"""
from __future__ import annotations

import os
import random
import warnings
from typing import Any, Sequence

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.collections import PolyCollection
from matplotlib.lines import Line2D
from sklearn.metrics import (
    accuracy_score,
    roc_auc_score,
    precision_recall_fscore_support,
    classification_report,
    confusion_matrix,
    roc_curve,
    auc,
    f1_score,
)
from sklearn.preprocessing import label_binarize

import torch
import torch.nn as nn
import torch.nn.functional as F

warnings.filterwarnings("ignore", category=FutureWarning)

# ============================================================
#  Visual settings (macaron academic style)
# ============================================================
sns.set_theme(style="ticks")
plt.rcParams["font.sans-serif"] = ["Arial", "Microsoft YaHei", "SimHei"]
plt.rcParams["axes.unicode_minus"] = False
plt.rcParams["figure.dpi"] = 120
plt.rcParams["savefig.dpi"] = 600

FONT_SIZE_LABEL = 10
FONT_SIZE_TICK = 9
FONT_SIZE_LEGD = 9

MACARON_LIGHT = ["#FFB7B2", "#FFDAC1", "#E2F0CB", "#B5EAD7", "#C7CEEA", "#E8C4EC", "#F3D1D1"]
MACARON_DARK = ["#E07A7A", "#E5A97E", "#B0C986", "#79C2A5", "#8FA0D2", "#B387B8", "#C29595"]

SEED = 42
np.random.seed(SEED)


# ============================================================
#  Random Seed
# ============================================================
def set_seed(seed: int = 1208) -> None:
    """Set all random seeds for reproducibility"""
    random.seed(seed)
    os.environ["PYTHONHASHSEED"] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def seed_worker(worker_id: int) -> None:
    """DataLoader worker seed setting"""
    worker_seed = torch.initial_seed() % (2**32)
    np.random.seed(worker_seed)
    random.seed(worker_seed)


# ============================================================
#  Loss Functions
# ============================================================
class FocalLoss(nn.Module):
    """Focal Loss for handling class imbalance"""

    def __init__(self, alpha: float = 1.0, gamma: float = 2.0, reduction: str = "mean"):
        super().__init__()
        self.alpha = alpha
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, logits: torch.Tensor, targets: torch.Tensor) -> torch.Tensor:
        ce_loss = F.cross_entropy(logits, targets, reduction="none")
        pt = torch.exp(-ce_loss)
        focal_loss = self.alpha * (1 - pt) ** self.gamma * ce_loss
        if self.reduction == "mean":
            return torch.mean(focal_loss)
        elif self.reduction == "sum":
            return torch.sum(focal_loss)
        return focal_loss


# ============================================================
#  Plotting Helpers (macaron academic style)
# ============================================================
def apply_unified_axis_style(ax, fontname='Arial'):
    for label in ax.get_xticklabels():
        label.set_fontname(fontname)
        label.set_fontsize(FONT_SIZE_TICK)
    for label in ax.get_yticklabels():
        label.set_fontname(fontname)
        label.set_fontsize(FONT_SIZE_TICK)


def remove_legend_title(legend):
    """彻底去除图例标题。"""
    if legend is not None:
        legend.set_title("")
        title = legend.get_title()
        if title is not None:
            title.set_text("")


# ============================================================
#  ROC Plotting (with Bootstrap 95% CI)
# ============================================================
def plot_multiclass_roc(
    all_labels: Sequence[int],
    all_logits: np.ndarray,
    class_names: list[str],
    save_path: str,
    n_bootstrap: int = 1000,
    seed: int = 42,
) -> None:
    """
    Plot multiclass ROC curves with micro/macro averages and 95% CI.
    Macaron academic style, no shadow.
    """
    all_labels = np.array(all_labels)
    all_logits = np.array(all_logits)
    C = all_logits.shape[1]
    N = len(all_labels)
    assert C == len(class_names), "class_names length must match logits columns"

    y_onehot = label_binarize(all_labels, classes=list(range(C)))
    if y_onehot.shape[1] != C:
        tmp = np.zeros((y_onehot.shape[0], C))
        tmp[:, : y_onehot.shape[1]] = y_onehot
        y_onehot = tmp

    # ---- Compute ROC ----
    fpr, tpr, roc_auc = {}, {}, {}
    for i in range(C):
        try:
            fpr[i], tpr[i], _ = roc_curve(y_onehot[:, i], all_logits[:, i])
            roc_auc[i] = auc(fpr[i], tpr[i])
        except Exception:
            fpr[i], tpr[i] = np.array([0, 1]), np.array([0, 1])
            roc_auc[i] = float("nan")

    try:
        fpr["micro"], tpr["micro"], _ = roc_curve(y_onehot.ravel(), all_logits.ravel())
        roc_auc["micro"] = auc(fpr["micro"], tpr["micro"])
    except Exception:
        fpr["micro"], tpr["micro"] = np.array([0, 1]), np.array([0, 1])
        roc_auc["micro"] = float("nan")

    all_fpr = np.unique(np.concatenate([fpr[i] for i in range(C)]))
    mean_tpr = np.concatenate(
        [np.interp(all_fpr, fpr[i], tpr[i]).reshape(-1, 1) for i in range(C)], axis=1
    ).mean(1)
    fpr["macro"], tpr["macro"] = all_fpr, mean_tpr
    try:
        roc_auc["macro"] = auc(fpr["macro"], tpr["macro"])
    except Exception:
        roc_auc["macro"] = float("nan")

    # ---- Bootstrap 95% CI ----
    rng = np.random.RandomState(seed)
    boot_aucs: dict[str, list[float]] = {i: [] for i in range(C)}
    boot_aucs["micro"] = []
    boot_aucs["macro"] = []

    for _ in range(n_bootstrap):
        idx = rng.randint(0, N, N)
        y_b = y_onehot[idx]
        logits_b = all_logits[idx]

        for i in range(C):
            try:
                boot_aucs[i].append(roc_auc_score(y_b[:, i], logits_b[:, i]))
            except Exception:
                boot_aucs[i].append(np.nan)

        try:
            boot_aucs["micro"].append(roc_auc_score(y_b.ravel(), logits_b.ravel()))
        except Exception:
            boot_aucs["micro"].append(np.nan)

        per_class = []
        for i in range(C):
            try:
                per_class.append(roc_auc_score(y_b[:, i], logits_b[:, i]))
            except Exception:
                per_class.append(np.nan)
        per_class_arr = np.array(per_class)
        boot_aucs["macro"].append(
            np.nanmean(per_class_arr) if not np.all(np.isnan(per_class_arr)) else np.nan
        )

    ci_dict = {}
    for k, vals in boot_aucs.items():
        vals_arr = np.array(vals)
        vals_arr = vals_arr[~np.isnan(vals_arr)]
        if len(vals_arr) == 0:
            ci_dict[k] = (float("nan"), float("nan"))
        else:
            ci_dict[k] = (np.percentile(vals_arr, 2.5), np.percentile(vals_arr, 97.5))

    # ---- Plot (macaron academic style, no shadow) ----
    plt.figure(figsize=(5, 4))

    def _plot(f, t, auc_val, ci_val, label, linewidth=1.0, linestyle='-'):
        if ci_val and not np.isnan(ci_val[0]):
            plt.plot(
                f, t, lw=linewidth, linestyle=linestyle,
                label=f"{label} (AUC={auc_val:.3f} [{ci_val[0]:.3f}-{ci_val[1]:.3f}])"
            )
        else:
            plt.plot(f, t, lw=linewidth, linestyle=linestyle, label=f"{label} (AUC={auc_val:.3f})")

    if not np.isnan(roc_auc.get("micro", np.nan)):
        _plot(fpr["micro"], tpr["micro"], roc_auc["micro"], ci_dict.get("micro"), "micro-average", linewidth=2.5)
    if not np.isnan(roc_auc.get("macro", np.nan)):
        _plot(fpr["macro"], tpr["macro"], roc_auc["macro"], ci_dict.get("macro"), "macro-average", linewidth=2)

    for i in range(C):
        if not np.isnan(roc_auc[i]):
            _plot(fpr[i], tpr[i], roc_auc[i], ci_dict.get(i), class_names[i], linewidth=1.2)

    plt.xlabel("False Positive Rate", fontsize=12)
    plt.ylabel("True Positive Rate", fontsize=12)
    leg = plt.legend(loc="lower right", fontsize='small', framealpha=0.9)
    remove_legend_title(leg)
    plt.grid(False)
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.close()


# ============================================================
#  Confusion Matrix with Precision/Recall (macaron academic style)
# ============================================================
def plot_confusion_matrix(
    all_labels: Sequence[int],
    all_preds: Sequence[int],
    class_names: list[str],
    save_path: str,
) -> None:
    cm = confusion_matrix(all_labels, all_preds, labels=list(range(len(class_names))))
    cm = np.array(cm)
    n = cm.shape[0]

    row_sum = cm.sum(axis=1)
    col_sum = cm.sum(axis=0)

    recall = np.divide(np.diag(cm), row_sum, out=np.zeros(n, dtype=float), where=row_sum != 0)
    precision = np.divide(np.diag(cm), col_sum, out=np.zeros(n, dtype=float), where=col_sum != 0)

    fig, ax = plt.subplots(figsize=(5, 5))
    ax.imshow(cm, cmap="Blues", aspect="equal", vmin=0)

    y_offset = 0.0

    for i in range(n):
        for j in range(n):
            color = "white" if cm[i, j] > cm.max() * 0.5 else "black"
            ax.text(j, i - y_offset, str(cm[i, j]), ha="center", va="center", fontsize=16, color=color,
                    fontname="Arial")

    x_labels = [f"{class_names[i]}" for i in range(n)]
    y_labels = [f"{class_names[i]}" for i in range(n)]

    ax.set_xticks(np.arange(n))
    ax.set_yticks(np.arange(n))

    ax.set_xticklabels(x_labels, rotation=45, ha="right", rotation_mode="anchor", fontsize=14, fontname="Arial")
    ax.set_yticklabels(y_labels, fontsize=14, fontname="Arial")

    for j in range(n):
        if col_sum[j] > 0:
            ax.text(
                j, -0.75, f"{precision[j] * 100:.1f}%",
                ha="center", va="center", rotation=45,
                fontsize=14, fontname="Arial"
            )

    recall_x_pos = n - 0.45 if n >= 3 else n - 0.48
    for i in range(n):
        if row_sum[i] > 0:
            ax.text(recall_x_pos, i - y_offset, f"{recall[i] * 100:.1f}%", ha="left", va="center", fontsize=14,
                fontname="Arial")

    ax.set_xlabel("Predicted labels", fontsize=14, fontname="Arial", labelpad=20)
    ax.set_ylabel("True labels", fontsize=14, fontname="Arial", labelpad=15)
    ax.xaxis.set_label_coords(0.45, -0.15)
    ax.yaxis.set_label_coords(-0.15, 0.45)
    for spine in ax.spines.values():
        spine.set_visible(False)

    right_margin = 0.6 if n >= 3 else 0.4
    ax.set_xlim(-0.6, n - 1 + right_margin + 0.5)
    ax.set_ylim(n - 0.5, -0.9)

    plt.tight_layout()
    plt.savefig(save_path, dpi=600, bbox_inches="tight")
    plt.close()


# ============================================================
#  AUC Boxplot (macaron academic style)
# ============================================================
def plot_auc_boxplot(
    all_labels: Sequence[int],
    all_logits: np.ndarray,
    class_names: list[str],
    save_path: str,
    n_bootstrap: int = 1000,
    seed: int = 42,
) -> None:
    """Plot AUC boxplot with bootstrap distribution per class."""
    all_labels = np.array(all_labels)
    all_logits = np.array(all_logits)
    C = all_logits.shape[1]
    N = len(all_labels)

    y_onehot = label_binarize(all_labels, classes=list(range(C)))
    if y_onehot.shape[1] != C:
        tmp = np.zeros((y_onehot.shape[0], C))
        tmp[:, : y_onehot.shape[1]] = y_onehot
        y_onehot = tmp

    rng = np.random.RandomState(seed)
    boot_aucs = {i: [] for i in range(C)}
    for _ in range(n_bootstrap):
        idx = rng.randint(0, N, N)
        for i in range(C):
            try:
                boot_aucs[i].append(roc_auc_score(y_onehot[idx][:, i], all_logits[idx][:, i]))
            except Exception:
                pass

    plt.figure(figsize=(5, 4))
    ax = plt.gca()
    for i in range(C):
        vals = np.array(boot_aucs[i])
        vals = vals[~np.isnan(vals)]
        if len(vals) == 0:
            continue
        t_light = MACARON_LIGHT[i % len(MACARON_LIGHT)]
        t_dark = MACARON_DARK[i % len(MACARON_DARK)]
        bp = ax.boxplot(vals, positions=[i], widths=0.35, showcaps=True, showfliers=False, patch_artist=True)
        for box in bp['boxes']:
            box.set_facecolor(t_light)
            box.set_alpha(0.65)
            box.set_edgecolor(t_dark)
            box.set_linewidth(0.9)
        for whisker in bp['whiskers']:
            whisker.set_color(t_dark)
            whisker.set_linewidth(0.9)
        for cap in bp['caps']:
            cap.set_color(t_dark)
            cap.set_linewidth(0.9)
        for median in bp['medians']:
            median.set_color(t_dark)
            median.set_linewidth(1.5)

    ax.set_xticks(range(C))
    ax.set_xticklabels(class_names)
    plt.xlabel("True labels", fontsize=FONT_SIZE_LABEL, labelpad=6, fontname="Arial")
    plt.ylabel("AUC Value", fontsize=FONT_SIZE_LABEL, labelpad=6, fontname="Arial")
    plt.ylim(0.68, 1.02)
    plt.yticks(np.arange(0.7, 1.01, 0.05), [f"{x:.2f}" for x in np.arange(0.7, 1.01, 0.05)])
    plt.grid(axis="y", linestyle=":", alpha=0.4, color="#B0B0B0")
    apply_unified_axis_style(ax, fontname="Arial")
    sns.despine()
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.close()


# ============================================================
#  F1 Violin Plot (macaron academic style)
# ============================================================
def plot_f1_violin(
    all_labels: Sequence[int],
    all_preds: Sequence[int],
    class_names: list[str],
    save_path: str,
    n_bootstrap: int = 1000,
    seed: int = 42,
) -> None:
    """Plot F1 violin plot with bootstrap distribution per class."""
    all_labels = np.array(all_labels)
    all_preds = np.array(all_preds)
    C = len(class_names)
    N = len(all_labels)

    rng = np.random.RandomState(seed)
    boot_f1s = {i: [] for i in range(C)}
    for _ in range(n_bootstrap):
        idx = rng.randint(0, N, N)
        boot_true = all_labels[idx]
        boot_pred = all_preds[idx]
        for i in range(C):
            true_c = (boot_true == i).astype(int)
            pred_c = (boot_pred == i).astype(int)
            boot_f1s[i].append(f1_score(true_c, pred_c, zero_division=0))

    rows = []
    for i in range(C):
        for v in boot_f1s[i]:
            rows.append({"Class": class_names[i], "F1": v})
    df_f1 = pd.DataFrame(rows)

    legend_elements = [
        Line2D([0], [0], marker='_', color='w', markeredgecolor='#555555', markeredgewidth=1.8, markersize=10,
               label='Median'),
        Line2D([0], [0], marker='D', color='w', markerfacecolor='#555555', markeredgecolor='#555555', alpha=0.6,
               markersize=5, label='Mean')
    ]

    plt.figure(figsize=(5, 4))
    palette = {cls: MACARON_LIGHT[i % len(MACARON_LIGHT)] for i, cls in enumerate(class_names)}
    ax = sns.violinplot(data=df_f1, x="Class", y="F1", order=class_names,
                        inner=None, cut=0, linewidth=0.8, palette=palette, saturation=1.0)

    violin_bodies = [c for c in ax.collections if isinstance(c, PolyCollection)]
    for idx, body in enumerate(violin_bodies[:C]):
        body.set_alpha(0.65)
        body.set_edgecolor(MACARON_DARK[idx % len(MACARON_DARK)])
        body.set_linewidth(0.9)

    means = df_f1.groupby("Class")["F1"].mean().reindex(class_names).values
    medians = df_f1.groupby("Class")["F1"].median().reindex(class_names).values
    for i in range(C):
        c_dark = MACARON_DARK[i % len(MACARON_DARK)]
        ax.plot(i, medians[i], marker='_', color=c_dark, markeredgewidth=1.8, markersize=10, zorder=3)
        ax.plot(i, means[i], marker='D', color=c_dark, markerfacecolor=c_dark, markersize=4, alpha=0.4, zorder=4)

    plt.xlabel("True labels", fontsize=FONT_SIZE_LABEL, labelpad=6, fontname="Arial")
    plt.ylabel("F1-Score", fontsize=FONT_SIZE_LABEL, labelpad=6, fontname="Arial")
    plt.ylim(-0.02, 1.02)
    plt.yticks(np.arange(0, 1.01, 0.2))
    plt.grid(axis="y", linestyle=":", alpha=0.4, color="#B0B0B0")
    apply_unified_axis_style(ax, fontname="Arial")
    sns.despine()
    leg = plt.legend(handles=legend_elements, loc="lower left", frameon=True, fontsize=FONT_SIZE_LEGD)
    remove_legend_title(leg)
    plt.tight_layout()
    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.close()


# ============================================================
#  Evaluation
# ============================================================
def evaluate(
    model: nn.Module,
    loader: Any,
    device: torch.device,
    num_classes: int,
    class_names: list[str] | None = None,
    save_csv: str | None = None,
    dataset: Any = None,
) -> dict[str, Any]:
    """
    Unified evaluation function, returns dict with all metrics and raw predictions.

    Returns
    -------
    dict with keys: acc, auc, f1, cm, labels, preds, probs, pids
    """
    model.eval()
    all_labels, all_preds, all_probs, all_pids = [], [], [], []

    with torch.no_grad():
        for batch_idx, batch in enumerate(loader):
            # Support different DataLoader output formats
            if isinstance(batch, dict):
                pids = batch.get("pid", None)
                inputs = {k: v.to(device) for k, v in batch.items() if k not in ("label", "pid")}
                labels = batch["label"].to(device)
                logits = model(**inputs)
            elif isinstance(batch, (list, tuple)) and len(batch) == 2:
                Xs, ys = batch
                if ys.dim() == 0:
                    ys = ys.unsqueeze(0)
                ys = ys.to(device)
                logits = model(Xs)
                labels = ys
                pids = None
            else:
                raise ValueError(f"Unsupported batch format: {type(batch)}")

            probs = F.softmax(logits, dim=1)
            preds = logits.argmax(dim=1)

            all_labels.extend(labels.cpu().tolist())
            all_preds.extend(preds.cpu().tolist())
            all_probs.extend(probs.cpu().numpy().tolist())

            if pids is not None:
                all_pids.extend(pids)
            elif dataset is not None and hasattr(dataset, "ids"):
                start = batch_idx * (loader.batch_size or 1)
                end = start + len(labels)
                all_pids.extend(dataset.ids[start:end])
            else:
                all_pids.extend([None] * len(labels))

    # ---- Metrics ----
    acc = accuracy_score(all_labels, all_preds) if all_labels else 0.0
    try:
        auc_score = roc_auc_score(all_labels, np.array(all_probs), multi_class="ovr")
    except Exception:
        auc_score = float("nan")
    _, _, f1, _ = precision_recall_fscore_support(
        all_labels, all_preds, average="macro", zero_division=0
    )
    cm = confusion_matrix(all_labels, all_preds, labels=list(range(num_classes)))

    # ---- Save predictions ----
    if save_csv and class_names:
        rows = []
        for pid, lab, pred, prob in zip(all_pids, all_labels, all_preds, all_probs):
            row = {"ID": pid, "true_label": int(lab), "predicted_label": int(pred)}
            for i in range(num_classes):
                row[f"class_{i}_prob"] = float(prob[i])
            rows.append(row)
        result_df = pd.DataFrame(rows)
        if save_csv.endswith(".xlsx"):
            result_df.to_excel(save_csv, index=False)
        else:
            result_df.to_csv(save_csv, index=False, encoding="utf-8-sig")

    return {
        "acc": acc,
        "auc": auc_score,
        "f1": f1,
        "cm": cm,
        "labels": all_labels,
        "preds": all_preds,
        "probs": all_probs,
        "pids": all_pids,
    }


# ============================================================
#  Summary Excel (per-class + global metrics with bootstrap CI)
# ============================================================
def calculate_detailed_metrics(y_true_bin, y_pred_bin, y_prob):
    """Calculate per-class detailed metrics (binary one-vs-rest)."""
    metrics = {}
    cm_bin = confusion_matrix(y_true_bin, y_pred_bin, labels=[0, 1])
    tn, fp, fn, tp = cm_bin.ravel()

    metrics['Sensitivity'] = tp / (tp + fn) if (tp + fn) > 0 else np.nan
    metrics['Specificity'] = tn / (tn + fp) if (tn + fp) > 0 else np.nan
    metrics['PPV (Precision)'] = tp / (tp + fp) if (tp + fp) > 0 else np.nan
    metrics['NPV'] = tn / (tn + fn) if (tn + fn) > 0 else np.nan
    metrics['F1-Score'] = f1_score(y_true_bin, y_pred_bin, zero_division=0)

    try:
        if len(np.unique(y_true_bin)) < 2:
            metrics['AUC'] = np.nan
        else:
            metrics['AUC'] = roc_auc_score(y_true_bin, y_prob)
    except Exception:
        metrics['AUC'] = np.nan

    return metrics


def save_summary_xlsx(
    output_path,
    class_names,
    y_true,
    y_pred,
    y_score,
    n_bootstrap=1000,
    seed=42,
    title_text="Classification Results",
):
    """
    Save summary Excel with per-class and global metrics (point estimate [95% CI]).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_score = np.array(y_score)
    n_classes = len(class_names)

    def point_ci_text(point, boot_vals):
        try:
            point = float(point)
        except Exception:
            return ""
        if np.isnan(point):
            return ""
        clean = [v for v in boot_vals if not np.isnan(v)]
        if len(clean) == 0:
            return f"{point:.4f}"
        low = np.percentile(clean, 2.5)
        high = np.percentile(clean, 97.5)
        return f"{point:.4f} [{low:.4f}-{high:.4f}]"

    # ---- Raw metrics ----
    raw_class = {}
    for i in range(n_classes):
        raw_class[i] = calculate_detailed_metrics(
            (y_true == i).astype(int), (y_pred == i).astype(int), y_score[:, i]
        )

    raw_acc = accuracy_score(y_true, y_pred)
    try:
        raw_macro_auc = roc_auc_score(
            label_binarize(y_true, classes=list(range(n_classes))),
            y_score, average='macro', multi_class='ovr'
        )
    except Exception:
        raw_macro_auc = np.nan
    try:
        raw_micro_auc = roc_auc_score(
            label_binarize(y_true, classes=list(range(n_classes))).ravel(),
            y_score.ravel()
        )
    except Exception:
        raw_micro_auc = np.nan
    raw_macro_f1 = f1_score(y_true, y_pred, average='macro', zero_division=0)
    raw_weighted_f1 = f1_score(y_true, y_pred, average='weighted', zero_division=0)

    # ---- Bootstrap ----
    rng = np.random.RandomState(seed)
    n = len(y_true)
    y_onehot = label_binarize(y_true, classes=list(range(n_classes)))

    boot_class = {i: {k: [] for k in ['Sensitivity', 'Specificity', 'PPV (Precision)', 'NPV', 'F1-Score', 'AUC']} for i in range(n_classes)}
    boot_acc, boot_macro_auc, boot_micro_auc, boot_macro_f1, boot_weighted_f1 = [], [], [], [], []

    for _ in range(n_bootstrap):
        idx = rng.randint(0, n, n)
        bt, bp, bs = y_true[idx], y_pred[idx], y_score[idx]

        boot_acc.append(accuracy_score(bt, bp))
        boot_macro_f1.append(f1_score(bt, bp, average='macro', zero_division=0))
        boot_weighted_f1.append(f1_score(bt, bp, average='weighted', zero_division=0))

        try:
            boot_micro_auc.append(roc_auc_score(y_onehot[idx].ravel(), bs.ravel()))
        except Exception:
            boot_micro_auc.append(np.nan)

        per_class = []
        for i in range(n_classes):
            m = calculate_detailed_metrics((bt == i).astype(int), (bp == i).astype(int), bs[:, i])
            for k in boot_class[i]:
                boot_class[i][k].append(m[k])
            if not np.isnan(m['AUC']):
                per_class.append(m['AUC'])
        boot_macro_auc.append(np.nanmean(per_class) if per_class else np.nan)

    # ---- Write Excel: rows=Metrics, columns=Classes ----
    wb = Workbook()
    ws = wb.active
    ws.title = "具体数据"

    dark_blue, mid_blue, light_grid = "2F5597", "8EA9DB", "D9E2F3"
    title_fill = PatternFill("solid", fgColor=dark_blue)
    header_fill = PatternFill("solid", fgColor=mid_blue)
    left_fill = PatternFill("solid", fgColor=mid_blue)
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10, color="000000")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=light_grid)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    metric_rows = ["Sensitivity", "PPV", "AUC", "F1-score", "Specificity", "NPV"]
    metric_keys = ["Sensitivity", "PPV (Precision)", "AUC", "F1-Score", "Specificity", "NPV"]
    headers = ["Metrics"] + list(class_names)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).value = title_text
    ws.cell(row=1, column=1).fill = title_fill
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = center

    # Header
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = white_font
        c.alignment = center
        c.border = border

    # Data rows: each row = metric, each column = class
    for row_i, (display_name, metric_key) in enumerate(zip(metric_rows, metric_keys)):
        row = 3 + row_i
        ws.cell(row=row, column=1, value=display_name)
        ws.cell(row=row, column=1).fill = left_fill
        ws.cell(row=row, column=1).font = white_font
        ws.cell(row=row, column=1).alignment = center

        for col_i, cls_name in enumerate(class_names):
            point = raw_class[col_i].get(metric_key, np.nan)
            vals = boot_class[col_i].get(metric_key, [])
            ws.cell(row=row, column=2 + col_i, value=point_ci_text(point, vals)).font = body_font
            ws.cell(row=row, column=2 + col_i).alignment = center
            ws.cell(row=row, column=2 + col_i).border = border

    # Format
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22
    for r in range(1, 3 + len(metric_rows) + 1):
        ws.row_dimensions[r].height = 24

    wb.save(output_path)
    print(f"[OK] Summary Excel saved to: {output_path}")
